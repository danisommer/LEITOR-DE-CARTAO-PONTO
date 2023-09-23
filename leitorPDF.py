import PyPDF2
import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

# Função para extrair horários de entrada e saída de um texto
def extrair_horarios(texto):
    padrao_horario = r"\d{2}:\d{2}"
    horarios = re.findall(padrao_horario, texto)
    return horarios

# Função para extrair todas as datas de uma página
def extrair_datas_da_pagina(page_text):
    padrao_data = r"\d{2}/\d{2}/\d{4}"
    datas = re.findall(padrao_data, page_text)
    return datas

# Função para ler o PDF e extrair os horários
def extrair_horarios_de_pdf(pdf_path):
    horarios_por_dia = {}
    
    pdf_file = open(pdf_path, "rb")
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    datas = []
    horarios = []
    dif = 0  # Inicialize a diferença como zero
    
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        page_text = page.extract_text()
        
        datas_da_pagina = extrair_datas_da_pagina(page_text)
        horarios_da_pagina = extrair_horarios(page_text)

        dif += len(datas_da_pagina) - len(horarios_da_pagina)
        
        if dif > 0:
            for _ in range(dif):
                horarios.append(("",""))  # Adicione um par de horários em branco
        elif dif < 0:
            for _ in range(abs(dif)):
                datas.append("")  # Adicione uma data em branco
    
        for data in datas_da_pagina:
            if horarios_da_pagina:  # Verifique se há horários disponíveis
                entrada = horarios_da_pagina.pop(0)
                saida = horarios_da_pagina.pop(0) if horarios_da_pagina else ""
            else:
                entrada, saida = "", ""  # Se não houver horários, use strings vazias
            
            if data not in horarios_por_dia:
                horarios_por_dia[data] = []
            
            horarios_por_dia[data].append((entrada, saida))
    
    pdf_file.close()
    
    return horarios_por_dia


# Função para salvar os horários em um arquivo Excel
def salvar_horarios_em_excel(horarios, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horários"
    
    # Adicionar cabeçalhos
    ws['A1'] = "Data"
    ws['B1'] = "Entrada"
    ws['C1'] = "Saída"
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    
    # Preencher dados
    row_num = 2
    for data, lista_horarios in horarios.items():
        for entrada, saida in lista_horarios:
            ws.cell(row=row_num, column=1, value=data)
            # Formatar entrada e saída como HH:MM (sem segundos)
            if entrada:
                entrada_time = datetime.strptime(entrada, "%H:%M").strftime("%H:%M")
            else:
                entrada_time = ""
            ws.cell(row=row_num, column=2, value=entrada_time)
            if saida:
                saida_time = datetime.strptime(saida, "%H:%M").strftime("%H:%M")
            else:
                saida_time = ""
            ws.cell(row=row_num, column=3, value=saida_time)
            row_num += 1
        
        # Verifique se ainda há horários na lista e preencha células em branco, se necessário
        while row_num < len(lista_horarios) + 2:
            ws.cell(row=row_num, column=2, value="")
            ws.cell(row=row_num, column=3, value="")
            row_num += 1
    
    # Salvar arquivo Excel
    wb.save(excel_path)

def selecionar_arquivo_pdf():
    pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if pdf_path:
        return pdf_path
    else:
        return None

def encontrar_nome_unico(nome_base, extensao):
    contador = 1
    while True:
        novo_nome = f"{nome_base}_{contador}.{extensao}"
        if not os.path.exists(novo_nome):
            return novo_nome
        contador += 1

def selecionar_diretorio_saida():
    diretorio_saida = filedialog.askdirectory()
    if diretorio_saida:
        return diretorio_saida
    else:
        return None

def processar_pdf():
    pdf_path = selecionar_arquivo_pdf()
    
    if pdf_path:
        excel_path_base = "horarios"  # Nome base do arquivo Excel de saída
        extensao_excel = "xlsx"  # Extensão do arquivo Excel
        
        diretorio_saida = selecionar_diretorio_saida()
        if not diretorio_saida:
            return

        if os.path.exists(pdf_path):
            messagebox.showinfo("Sucesso", "Leitura bem sucedida!")
        else:
            messagebox.showerror("Erro", "Erro ao abrir o arquivo PDF. Certifique-se de deixá-lo na mesma pasta que o programa.")
            return

        horarios = extrair_horarios_de_pdf(pdf_path)
        excel_path = os.path.join(diretorio_saida, encontrar_nome_unico(excel_path_base, extensao_excel))

        salvar_horarios_em_excel(horarios, excel_path)

        if os.path.exists(excel_path):
            messagebox.showinfo("Sucesso", "Excel criado com sucesso em: " + excel_path)
        else:
            messagebox.showerror("Erro", "Erro, não foi possível criar o Excel.")

# Criação da interface gráfica
root = tk.Tk()
root.title("Extrair Horários de PDF")

frame = tk.Frame(root)
frame.pack(padx=80, pady=30)

label = tk.Label(frame, text="Selecione um arquivo PDF:")
label.pack()

process_button = tk.Button(frame, text="Processar PDF", command=processar_pdf)
process_button.pack()

exit_button = tk.Button(frame, text="Sair", command=root.quit)
exit_button.pack()

root.mainloop()
